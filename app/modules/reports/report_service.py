from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path
import pandas as pd
from sqlalchemy.orm import Session
from fastapi import HTTPException

# ReportLab imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# OpenPyXL imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.config import settings
from app.models.dataset import Dataset
from app.models.report import Report
from app.models.training_job import TrainingJob
from app.models.trained_model import TrainedModel
from app.modules.eda.eda_service import EDAService
from app.modules.explainability.shap_service import SHAPService


class ReportService:
    @staticmethod
    def generate_report_sync(db: Session, report_id: int) -> None:
        """
        Generates a PDF or Excel report for the specified report ID.
        Updates status to completed/failed and sets the storage path.
        """
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            return

        report.status = "generating"
        db.commit()

        try:
            dataset = db.query(Dataset).filter(Dataset.id == report.dataset_id).first()
            if not dataset:
                raise ValueError("Dataset not found")

            # Load cleaned dataset
            cleaned_path = settings.data_path / f"cleaned/{dataset.id}/cleaned.csv"
            if not cleaned_path.exists():
                raise ValueError("Cleaned dataset not found. Please clean dataset before reporting.")

            df = pd.read_csv(cleaned_path)
            metadata = dataset.get_schema_metadata()
            
            # Fetch latest training job to include leaderboard
            job = (
                db.query(TrainingJob)
                .filter(TrainingJob.dataset_id == dataset.id)
                .order_by(TrainingJob.completed_at.desc())
                .first()
            )

            # Ensure reports storage directory exists
            report_dir = settings.reports_dir / str(report.id)
            report_dir.mkdir(parents=True, exist_ok=True)

            if report.report_type == "pdf":
                relative_path = f"reports/{report.id}/report.pdf"
                absolute_path = settings.data_path / relative_path
                ReportService._build_pdf(absolute_path, dataset, df, metadata, job)
            elif report.report_type == "excel":
                relative_path = f"reports/{report.id}/report.xlsx"
                absolute_path = settings.data_path / relative_path
                ReportService._build_excel(absolute_path, dataset, df, metadata, job)
            else:
                raise ValueError(f"Unknown report type: {report.report_type}")

            report.storage_path = relative_path
            report.status = "completed"
            db.commit()

        except Exception as e:
            report.status = "failed"
            db.commit()
            print(f"Report generation failed for report {report_id}: {str(e)}")
            raise e

    @staticmethod
    def _build_pdf(
        path: Path,
        dataset: Dataset,
        df: pd.DataFrame,
        metadata: dict,
        job: TrainingJob | None
    ) -> None:
        """Assembles a beautiful, print-ready PDF report."""
        # Ensure static plots exist for the dataset
        plot_paths = EDAService.generate_static_plots(dataset.id, df)
        
        # If best model exists, try generating global SHAP summary plot if not already present
        shap_plot_path = settings.plots_dir / str(dataset.id) / "shap_summary.png"
        if job and job.status == "completed" and not shap_plot_path.exists():
            best_model = next((m for m in job.models if m.is_best), None)
            if best_model:
                try:
                    # We can load model payload and generate plot
                    from app.modules.prediction.predict_service import PredictionService
                    payload = PredictionService.load_model_payload(best_model.id)
                    SHAPService.generate_global_importance(dataset.id, payload, df)
                except Exception as e:
                    print(f"Failed to generate SHAP plot for PDF: {str(e)}")

        doc = SimpleDocTemplate(str(path), pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        story = []

        styles = getSampleStyleSheet()
        
        # Premium custom styles
        title_style = ParagraphStyle(
            "CoverTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=24,
            leading=28,
            textColor=colors.HexColor("#1e3a8a"),
            alignment=0, # Left-aligned
            spaceAfter=10
        )
        subtitle_style = ParagraphStyle(
            "CoverSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=12,
            leading=16,
            textColor=colors.HexColor("#4b5563"),
            spaceAfter=30
        )
        h1_style = ParagraphStyle(
            "SectionH1",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#1e3a8a"),
            spaceBefore=15,
            spaceAfter=10,
            keepWithNext=True
        )
        body_style = ParagraphStyle(
            "ReportBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#1f2937"),
            spaceAfter=8
        )

        # Cover Page
        story.append(Paragraph("AutoInsight Platform Report", title_style))
        story.append(Paragraph(f"Dataset Analysis & Machine Learning Overview", subtitle_style))
        story.append(Spacer(1, 10))

        # Metadata Table
        meta_data = [
            [Paragraph("<b>Filename</b>", body_style), Paragraph(dataset.filename, body_style)],
            [Paragraph("<b>Uploaded At</b>", body_style), Paragraph(dataset.uploaded_at.strftime("%Y-%m-%d %H:%M UTC"), body_style)],
            [Paragraph("<b>Data Rows</b>", body_style), Paragraph(str(dataset.rows), body_style)],
            [Paragraph("<b>Data Columns</b>", body_style), Paragraph(str(dataset.columns), body_style)],
            [Paragraph("<b>Duplicates Dropped</b>", body_style), Paragraph(str(metadata.get("duplicates", 0)), body_style)],
        ]
        t_meta = Table(meta_data, colWidths=[150, 350])
        t_meta.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
            ("PADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 20))

        # Section 1: Data Profiling Summary
        story.append(Paragraph("1. Data Profiling & Quality", h1_style))
        story.append(Paragraph(
            "Below is a listing of the columns found in the dataset, along with their inferred data types "
            "and percentage of missing values.",
            body_style
        ))

        dtypes = metadata.get("dtypes", {})
        missing = metadata.get("missing_values", {})
        outliers = metadata.get("outlier_counts", {})

        profile_table_data = [
            ["Column Name", "Inferred Type", "Missing %", "Outlier Count"]
        ]
        for col in df.columns:
            profile_table_data.append([
                col,
                dtypes.get(col, "Unknown"),
                missing.get(col, "0%"),
                str(outliers.get(col, 0))
            ])
            
        t_profile = Table(profile_table_data, colWidths=[180, 110, 100, 110])
        t_profile.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t_profile)
        
        # Page break before plots
        story.append(PageBreak())

        # Section 2: EDA Visualizations
        story.append(Paragraph("2. Exploratory Data Analysis", h1_style))
        story.append(Paragraph(
            "This section highlights the statistical relationships and correlations within the numeric features "
            "of the dataset.",
            body_style
        ))
        
        heatmap_path = settings.plots_dir / str(dataset.id) / "correlation_heatmap.png"
        if heatmap_path.exists():
            story.append(Image(str(heatmap_path), width=450, height=340))
            story.append(Spacer(1, 10))
            story.append(Paragraph("<i>Figure 2.1: Pearson correlation heatmap among numeric features.</i>", ParagraphStyle("caption", parent=styles["Italic"], fontSize=8, alignment=1)))
        else:
            story.append(Paragraph("No numeric features available to plot correlation heatmap.", body_style))

        story.append(Spacer(1, 15))

        # Section 3: AutoML & Leaderboard
        story.append(Paragraph("3. AutoML Leaderboard", h1_style))
        if job:
            story.append(Paragraph(
                f"AutoML trained multiple candidate models targeting column <b>{job.target_column}</b> "
                f"(Problem Type: <i>{job.problem_type}</i>). The leaderboard ranking is shown below:",
                body_style
            ))

            # Fetch models sorted by metric
            models = sorted(job.models, key=lambda m: m.primary_metric_value or 0.0, reverse=True)

            primary_metric_name = "F1-Score" if job.problem_type == "classification" else "R²"
            leaderboard_data = [
                ["Rank", "Algorithm", primary_metric_name, "Best Model?"]
            ]
            for idx, m in enumerate(models):
                leaderboard_data.append([
                    str(idx + 1),
                    m.algorithm.replace("_", " ").title(),
                    f"{m.primary_metric_value:.4f}" if m.primary_metric_value is not None else "N/A",
                    "YES (Winner)" if m.is_best else "No"
                ])

            t_leaderboard = Table(leaderboard_data, colWidths=[60, 180, 130, 130])
            t_leaderboard.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(t_leaderboard)

            # Section 4: Explainability
            story.append(Spacer(1, 15))
            story.append(Paragraph("4. Explainable AI (SHAP)", h1_style))
            story.append(Paragraph(
                "SHAP (Shapley Additive exPlanations) values outline the contribution of each feature "
                "to the model's predictions. The plot below illustrates the global importances for the best model.",
                body_style
            ))

            if shap_plot_path.exists():
                story.append(Image(str(shap_plot_path), width=430, height=320))
                story.append(Spacer(1, 10))
                story.append(Paragraph("<i>Figure 4.1: SHAP global feature importances for the best model.</i>", ParagraphStyle("caption2", parent=styles["Italic"], fontSize=8, alignment=1)))
            else:
                story.append(Paragraph("SHAP summary plot is not available.", body_style))

        else:
            story.append(Paragraph("No model training jobs have been executed for this dataset yet.", body_style))

        doc.build(story)

    @staticmethod
    def _build_excel(
        path: Path,
        dataset: Dataset,
        df: pd.DataFrame,
        metadata: dict,
        job: TrainingJob | None
    ) -> None:
        """Generates a structured, multi-sheet Excel report."""
        wb = openpyxl.Workbook()
        
        # Sheet 1: Metadata & Overview
        ws_meta = wb.active
        ws_meta.title = "Metadata"
        
        # Styles
        title_font = Font(name="Arial", size=16, bold=True, color="1F497D")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Arial", size=10, bold=True)
        regular_font = Font(name="Arial", size=10)
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F5f8", end_color="F2F5f8", fill_type="solid")

        # Set Title
        ws_meta.cell(row=1, column=1, value="AutoInsight Overview").font = title_font
        ws_meta.cell(row=2, column=1, value="Report Generated on " + datetime.now().strftime("%Y-%m-%d %H:%M")).font = Font(italic=True, size=9)
        
        meta_rows = [
            ("Filename", dataset.filename),
            ("Uploaded At", dataset.uploaded_at.strftime("%Y-%m-%d %H:%M UTC")),
            ("Data Rows", dataset.rows),
            ("Data Columns", dataset.columns),
            ("Exact Duplicates Dropped", metadata.get("duplicates", 0)),
        ]
        
        ws_meta.cell(row=4, column=1, value="Property").font = header_font
        ws_meta.cell(row=4, column=1).fill = header_fill
        ws_meta.cell(row=4, column=2, value="Value").font = header_font
        ws_meta.cell(row=4, column=2).fill = header_fill
        
        for idx, (prop, val) in enumerate(meta_rows):
            r = idx + 5
            ws_meta.cell(row=r, column=1, value=prop).font = bold_font
            ws_meta.cell(row=r, column=2, value=val).font = regular_font
            if r % 2 == 1:
                ws_meta.cell(row=r, column=1).fill = zebra_fill
                ws_meta.cell(row=r, column=2).fill = zebra_fill
                
        # Sheet 2: Leaderboard
        ws_leader = wb.create_sheet(title="Leaderboard")
        ws_leader.cell(row=1, column=1, value="AutoML Model Leaderboard").font = title_font
        
        if job:
            ws_leader.cell(row=3, column=1, value="Target Column:").font = bold_font
            ws_leader.cell(row=3, column=2, value=job.target_column).font = regular_font
            ws_leader.cell(row=4, column=1, value="Problem Type:").font = bold_font
            ws_leader.cell(row=4, column=2, value=job.problem_type).font = regular_font

            headers = ["Rank", "Algorithm", "Primary Metric Value", "Is Best Model?"]
            for col_idx, text in enumerate(headers):
                cell = ws_leader.cell(row=6, column=col_idx+1, value=text)
                cell.font = header_font
                cell.fill = header_fill
                
            models = sorted(job.models, key=lambda m: m.primary_metric_value or 0.0, reverse=True)
            for idx, m in enumerate(models):
                r = idx + 7
                ws_leader.cell(row=r, column=1, value=idx+1).font = regular_font
                ws_leader.cell(row=r, column=2, value=m.algorithm.replace("_", " ").title()).font = regular_font
                ws_leader.cell(row=r, column=3, value=m.primary_metric_value).font = regular_font
                ws_leader.cell(row=r, column=4, value="YES" if m.is_best else "No").font = regular_font
                if r % 2 == 1:
                    for col_idx in range(1, 5):
                        ws_leader.cell(row=r, column=col_idx).fill = zebra_fill
        else:
            ws_leader.cell(row=3, column=1, value="No training runs executed yet.").font = regular_font

        # Sheet 3: Cleaned Data (up to 1000 rows to avoid excessive sheet sizes)
        ws_data = wb.create_sheet(title="Cleaned Data Sample")
        ws_data.cell(row=1, column=1, value="Cleaned Data (First 1000 rows)").font = title_font
        
        sample_df = df.head(1000)
        
        # Headers
        for col_idx, col_name in enumerate(sample_df.columns):
            cell = ws_data.cell(row=3, column=col_idx+1, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            
        # Rows
        for r_idx, row in enumerate(dataframe_to_rows(sample_df, index=False, header=False)):
            r = r_idx + 4
            for c_idx, val in enumerate(row):
                # Excel struggles with numpy types sometimes, parse properly
                if isinstance(val, (np.integer, np.int64)):
                    val = int(val)
                elif isinstance(val, (np.floating, np.float64)):
                    val = float(val)
                elif isinstance(val, np.bool_):
                    val = bool(val)
                
                cell = ws_data.cell(row=r, column=c_idx+1, value=val)
                cell.font = regular_font
                if r % 2 == 1:
                    cell.fill = zebra_fill
                    
        wb.save(str(path))
